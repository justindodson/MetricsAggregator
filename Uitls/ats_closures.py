import openpyxl
import tkinter as tk
from tkinter import filedialog
from utils import build_sheet_from_book, site_name_string_manager

class ATSClosures:

    """
        This class creates an object from the ats closures data
        built from the data in the Excel file provided.

        The actions of the class are to collect the appropriate data
        and perform calculations on the data for determining the closure
        rate of the ATS Findings by sites.

        The public endpoints will grant access to the data that is configured
        in a manner that is easily readable and understandable.
        The class will also export the data into a format to be used for building
        out plots in Excel w/ PPT

    """

    def __init__(self, file_path, *args, **kwargs):
        self.sites = []
        self.site_metrics = {}

        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            print()
            print("This may take a few seconds for large files.")
            print('Processing Data...')
            
            self.sheet = build_sheet_from_book(file_path)
            self.__append_to_sites_array(self.__populate_sites(self.sheet))
            metrics = self.__populate_site_metrics(self.sheet, self.sites)
            self.__merge_metrics(metrics)
        else:
            raise FileExistsError("Incorrect file format. Only .xlsx or .xls files accepted.")
        
    def __append_to_sites_array(self, sites):
        for site in sites:
            self.sites.append(site)

    def __merge_metrics(self, new_metrics_dict):
        self.site_metrics.update(new_metrics_dict)


    def __populate_sites(self, sheet):
        sites_array = []
        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=5).value is not None:
                if sheet.cell(row=i, column=5).value.startswith('3 US'):
                    site = site_name_string_manager(sheet.cell(row=i, column=5).value)

                    if site not in sites_array and site is not 'Dept' and site is not '' and site is not None:
                        sites_array.append(site)
        
        return sites_array

    def __populate_site_metrics(self, sheet, sites_array):
        metrics = {}
        if len(sites_array) > 0:

            for i in range(len(sites_array)):
                metrics[sites_array[i]] = {'total':0, "open":0, "closed":0, "open_past_due":0, 'closed_past_due':0}

                for j in range(1, sheet.max_row):
                    
                    if sheet.cell(row=j, column=6).value is not None and sheet.cell(row=j, column=5).value is not None:
                        closure_value = sheet.cell(row=j, column=6).value.strip()
                        site = site_name_string_manager(sheet.cell(row=j, column=5).value)
                    
                        if site == sites_array[i]:
                            metrics[site]['total'] += 1

                            if closure_value == 'Closed':
                                # check if past due
                                if self.__check_for_closed_past_due(j):
                                    metrics[site]['closed_past_due'] += 1
                                else:
                                    metrics[site]['closed'] += 1
                            elif closure_value == 'Open':
                                # check if closed past due
                                if self.__check_for_open_past_due(j):
                                    metrics[site]['open_past_due'] += 1
                                else:
                                    metrics[site]['open'] += 1
        else:
            raise ValueError("Cannot create Site Metrics With No Sites. Check to make sure sites are available")

        return metrics
    
    def __check_for_closed_past_due(self, row):
        days_past_closure = self.sheet.cell(row=row, column=8).value

        if days_past_closure is not None and not isinstance(days_past_closure, str):
            if days_past_closure > 0:
                return True
        return False


    def __check_for_open_past_due(self, row):
        days_to_close = self.sheet.cell(row=row, column=7).value

        if days_to_close is not None and not isinstance(days_to_close, str):
            if days_to_close > 30:
                return True
        return False

    def show_full_metrics(self):
        for site in self.sites:
            print('{} : \n\tTotal: {}\n\tOpen: {}\n\tOpen Past Due: {}\n\tClosed Past Due: {}\n\tClosed: {}'.format(site, self.site_metrics[site]['total'], 
            self.site_metrics[site]['open'], self.site_metrics[site]['open_past_due'], self.site_metrics[site]['closed_past_due'], self.site_metrics[site]['closed']))

    def calculate_closure_rate(self, site):
        total = int(self.site_metrics[site]['total'])
        cpd = int(self.site_metrics[site]['closed_past_due'])
        opd = int(self.site_metrics[site]['open_past_due'])
        
        return 100 - round((((cpd + opd) * 100) / total))

    def prep_for_table(self):
        data = []
        for site, metrics in self.site_metrics.items():
            row = []
            row.append(site)
            row.append(self.calculate_closure_rate(site))
            for i in metrics:
                row.append(self.site_metrics[site][i])
            data.append(row)

        return data

    def generate_excel_report(self):
        new_book = openpyxl.Workbook()
        ws = new_book.active
        metrics_table_data = self.prep_for_table()

        ws.append(['Site', 'Closure Rate', 'Total', 'Open', 'Closed', 'Open Past Due', 'Closed Past Due'])

        for row in metrics_table_data:
            ws.append(row)

        window = tk.Tk()
        window.withdraw()
        file_path = filedialog.asksaveasfilename()
        new_book.save(file_path + '.xlsx')

    def add_another_page(self):
        print('Select a File to add to the metric report')
        root = tk.Tk()
        root.withdraw()

        file_path = filedialog.askopenfilename()
        print()
        print("This may take a few seconds for large files.")
        print('Processing Data...')
        new_sheet = build_sheet_from_book(file_path)
        sites = self.__populate_sites(new_sheet)
        self.__append_to_sites_array(sites)

        metrics = self.__populate_site_metrics(new_sheet, sites)
        self.__merge_metrics(metrics)



root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()
comp = ATSClosures(file_path)

comp.add_another_page()
comp.add_another_page()
comp.add_another_page()
comp.add_another_page()

comp.generate_excel_report()
comp.show_full_metrics()