import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import numpy as np

def calculate_completion_rate(site):
    total = site_dict[site]['total']
    closed = site_dict[site]['closed']
    return (closed * 100) / total

def site_name_string_manager(site):
    return site[5:].split('(')[0].split('TX')[0]

def plot(site_dict):
    labels = []
    comp_rate = []
    cpd = []
    opd = []
    missed = []

    for k,v in site_dict.items():
        labels.append(k)
        comp_rate.append(calculate_completion_rate(k))
        cpd.append(v['closed_past_due'])
        opd.append(v['open_past_due'])
        missed.append(v['missed'])

    x = np.arange(len(labels))
    width = 0.35 #bar width

    fig, ax = plt.subplots()

    rects1 = ax.bar(x - width, comp_rate, width, label='Closure Rate (%)')
    rects2 = ax.bar(x - width/2, cpd, width, label = 'Closed Past Due')
    rects3 = ax.bar(x + width/2, opd, width, label='Open Past Due')
    rects4 = ax.bar(x + width, missed, width, label='Missed')

    ax.set_ylabel('Closure Rate')
    ax.set_title('Compliance Calendar Closures by Site')
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.legend()

    plt.setp(ax.get_xticklabels(), rotation=60, horizontalalignment='right')

    plt.show()


sites = []
site_dict = {}

wb = openpyxl.load_workbook("C:/Users/204055604/Desktop/compliance.xlsx")

sheet = wb.get_sheet_by_name('On-Shore Wind - NAM - OM and Se')

count = 0
for i in range(1, sheet.max_row):

    if site_name_string_manager(sheet.cell(row=i, column=1).value) not in sites and not 'Dept' in  sheet.cell(row=i, column=1).value:
        sites.append(site_name_string_manager(sheet.cell(row=i, column=1).value))

# the tuple layout is (total, closed, open, open_past_due, closed_past_due, missed )
for i in range(len(sites)):
    site_dict[sites[i]] = {'total':0, 'closed':0, 'open':0, 'open_past_due':0, 'closed_past_due':0, 'missed':0}

    for j in range(1, sheet.max_row):

        if site_name_string_manager(sheet.cell(row=j, column=1).value) == sites[i]:
            site_dict[sites[i]]['total'] = site_dict[sites[i]]['total'] + 1
            
            if sheet.cell(row=j, column=2).value.strip() == 'Closed':
                site_dict[sites[i]]['closed'] = site_dict[sites[i]]['closed'] + 1

            elif sheet.cell(row=j, column=2).value.strip() == 'Open':
                site_dict[sites[i]]['open'] = site_dict[sites[i]]['open'] + 1

            elif sheet.cell(row=j, column=2).value.strip() == 'Open Past Due':
                site_dict[sites[i]]['open_past_due'] = site_dict[sites[i]]['open_past_due'] + 1
            elif sheet.cell(row=j, column=2).value.strip() == 'Closed Past Due':
                site_dict[sites[i]]['closed_past_due'] = site_dict[sites[i]]['closed_past_due'] + 1

            elif sheet.cell(row=j, column=2).value.strip() == 'Missed':
                site_dict[sites[i]]['missed'] = site_dict[sites[i]]['missed'] + 1


for site in sites:
    rate = int(round(calculate_completion_rate(site)))
    print("{} - {}%".format(site,rate))

plot(site_dict)