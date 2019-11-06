import openpyxl
import tkinter as tk
from tkinter import filedialog
from Uitls.utils import build_sheet_from_book, site_name_string_manager


class ComplianceCalendar:

    def __init__(self, file_path_list, region, *args, **kwargs):
        self.sites = []
        self.site_metrics = {}
        self.region = region

        for file in file_path_list:
            self.__build_file_metric(file)

    """
        -- Private class functions for building the metric report --
    """
    # creates an excel sheet object and parses the sheet
    # to build add all sites to the site list and metric dict
    def __build_file_metric(self, file_path):
        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            sheet = build_sheet_from_book(file_path)
            self.__append_to_sites_array(self.__populate_sites(sheet))
            metrics = self.__populate_site_metrics(sheet, self.sites)
            self.__merge_metrics(metrics)
        else:
            raise FileExistsError("Incorrect file format. Only .xlsx or .xls files accepted")

    # appends a new site list to the self.sites list
    def __append_to_sites_array(self, sites):
        for site in sites:
            self.sites.append(site)

    # merges a new metrics dict to the self.metrics dict
    def __merge_metrics(self, new_metrics_dict):
        self.site_metrics.update(new_metrics_dict)

    # takes an excel sheet, parses the data and finds the rows that contain
    # sites from the specified region and builds and returns a list of sites.
    def __populate_sites(self, sheet):
        sites_array = []
        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=9).value is not None:
                if sheet.cell(row=i, column=9).value.startswith('{} US'.format(self.region)): # uses specified region to search
                    site_name = site_name_string_manager(sheet.cell(row=i, column=9).value)

                    if site_name not in sites_array and not 'Dept' in sheet.cell(row=i, column=9).value:
                        sites_array.append(site_name)
                # TODO: create another if statement to handle the Canada Sites
        return sites_array

    # Builds the site metrics dictionary from the sheet and sites array
    def __populate_site_metrics(self, sheet, sites_array):
        metrics = {}
        if len(sites_array) > 0:

            # for each site, build the metrics dict.
            for i in range(len(sites_array)):
                metrics[sites_array[i]] = {'total': 0, 'closed': 0, 'open_past_due': 0, 'closed_past_due': 0,
                                           'missed': 0, 'open': 0}

                for j in range(1, sheet.max_row):

                    if sheet.cell(row=j, column=9).value is not None:
                        closure_value = sheet.cell(row=j, column=15).value.strip()

                        if site_name_string_manager(sheet.cell(row=j, column=9).value) == sites_array[i]:
                            metrics[sites_array[i]]['total'] += 1

                            if closure_value == 'Closed':
                                metrics[sites_array[i]]['closed'] += 1

                            elif closure_value == 'Open Past Due':
                                metrics[sites_array[i]]['open_past_due'] += 1

                            elif closure_value == 'Closed Past Due':
                                metrics[sites_array[i]]['closed_past_due'] += 1

                            elif closure_value == 'Missed':
                                metrics[sites_array[i]]['missed'] += 1

                            elif closure_value == 'Open':
                                metrics[sites_array[i]]['open'] += 1

        return metrics

    """
        -- Metric and Export public functions -- 
    """

    # calculates the closure rates for each site.
    def __calculate_closure_rates(self, site):
        total = self.site_metrics[site]['total'] - self.site_metrics[site]['open']
        closed = self.site_metrics[site]['closed']
        return int(round((closed * 100) / total, 2))

    # method to show closure rates for all sites
    # TODO: Convert this to a function to send to view
    def show_closures(self):
        for site in self.sites:
            rate = int(round(self.__calculate_closure_rates(site), 2))
            print("{} - {}%".format(site, rate))

    # Creates the full metric view for all sites
    # TODO: Convert this to a function to send to view
    def view_full_metric_report(self):
        for site in self.sites:
            print(
                '{} : \n\tClosure Rate: {}% \n\tTotal: {} \n\tClosed: {} \n\tOpen: {} \n\tOpen Past Due: {} \n\tClosed Past Due: {} \n\tMissed: {}'
                .format(site, self.__calculate_closure_rates(site), self.site_metrics[site]['total'],
                        self.site_metrics[site]['closed'], self.site_metrics[site]['open']
                        , self.site_metrics[site]['open_past_due'], self.site_metrics[site]['closed_past_due'],
                        self.site_metrics[site]['missed']))
            print()

    # generates a metric for the entire region
    # TODO: Convert this to a function to send to view
    def generate_region_closure_rate(self):
        # refactor to utilize a global file that is created as the final
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename()
        ws = build_sheet_from_book(file_path)

        metric_list = []
        total = 0
        total_open = 0
        closed = 0
        total_closed_past_due = 0
        total_open_past_due = 0
        total_missed = 0

        for i in range(2, ws.max_row):
            total += ws.cell(row=i, column=3).value
            closed += ws.cell(row=i, column=4).value
            total_open_past_due += ws.cell(row=i, column=5).value
            total_closed_past_due += ws.cell(row=i, column=6).value
            total_missed += ws.cell(row=i, column=7).value
            total_open += ws.cell(row=i, column=8).value

        total = total - total_open
        metric_list.append(total)
        metric_list.append(closed)
        metric_list.append(total_open_past_due)
        metric_list.append(total_closed_past_due)
        metric_list.append(total_missed)

        return metric_list

    # builds metrics in a form for a pie chart
    # TODO: Convert this to a function to send to view
    def calculate_pie_chart_metrics(metric_totals_list):
        rates = []
        closed_rate = round((metric_totals_list[1] * 100 / metric_totals_list[0]), 2)
        opd_rate = round((metric_totals_list[2] * 100 / metric_totals_list[0]), 2)
        cpd_rate = round((metric_totals_list[3] * 100 / metric_totals_list[0]), 2)
        missed_rate = round((metric_totals_list[4] * 100 / metric_totals_list[0]), 2)

        rates.append('')
        rates.append(closed_rate)
        rates.append(opd_rate)
        rates.append(cpd_rate)
        rates.append(missed_rate)

        new_book = openpyxl.Workbook()
        sheet = new_book.active

        sheet.append(['', 'Closed', 'Open Past Due', 'Closed Past Due', 'Missed'])
        sheet.append(rates)
        sheet.append(['Totals', 'Closed', 'Open Past Due', 'Closed Past Due', 'Missed'])
        sheet.append(metric_totals_list)
        window = tk.Tk()
        window.withdraw()
        file_path = filedialog.asksaveasfilename()

        new_book.save(file_path + '.xlsx')

    # Builds a table out of the metrics in a format that is easier used
    # the format is a list of lists where each inner list will be appended
    # to a separate row in the correct order to the final sheet.
    def prep_for_table(self):
        data = []
        for site, metrics in self.site_metrics.items():
            row = []
            row.append(site)
            row.append(self.__calculate_closure_rates(site))
            for i in metrics:
                row.append(self.site_metrics[site][i])
            data.append(row)

        return data

    # Builds the final excel report
    # TODO: Convert this to a function to send to view
    def generate_excel_report(self, file_path):
        # TODO: Add totals to report
        new_book = openpyxl.Workbook()
        ws = new_book.active
        metrics_table_data = self.prep_for_table()

        ws.append(['Site', 'Closure Rate', 'Total', 'Closed', 'Open Past Due', 'Closed Past Due', 'Missed', 'Open'])

        for row in metrics_table_data:
            ws.append(row)

        new_book.save(file_path + '.xlsx')
